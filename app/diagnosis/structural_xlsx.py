"""Structural analysis for XLSX files using openpyxl."""

from __future__ import annotations

import asyncio

from app.core.config import settings
from app.core.log import logger
from app.schema.diagnosis import (
    DiagnosisIssue,
    IssueSeverity,
    IssueSource,
    IssueType,
)

__all__ = ("analyze_xlsx",)

_SAFE_FONTS = {
    "Arial", "Calibri", "Cambria", "Courier New", "Georgia",
    "Helvetica", "Tahoma", "Times New Roman", "Verdana",
    "Consolas", "Segoe UI", "Trebuchet MS",
}

# Standard paper sizes in inches (width, height) for portrait orientation
_PAPER_SIZES: dict[int, tuple[float, float]] = {
    1: (8.5, 11.0),    # Letter
    9: (8.27, 11.69),   # A4
}


async def analyze_xlsx(file_path: str, job_id: str) -> list[DiagnosisIssue]:
    """Perform structural analysis on an XLSX file."""
    logger.info(f"Job {job_id}: running XLSX structural analysis on {file_path}")
    return await asyncio.to_thread(_analyze_xlsx_sync, file_path, job_id)


def _analyze_xlsx_sync(file_path: str, job_id: str) -> list[DiagnosisIssue]:
    from openpyxl import load_workbook

    issues: list[DiagnosisIssue] = []
    try:
        wb = load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            issues.extend(_check_page_setup(ws, sheet_name))
            issues.extend(_check_margins(ws, sheet_name))
            issues.extend(_check_column_widths(ws, sheet_name))
            issues.extend(_check_fonts(ws, sheet_name))
            issues.extend(_check_print_area(ws, sheet_name))
        wb.close()
    except Exception:
        logger.exception(f"Job {job_id}: XLSX structural analysis failed")
    return issues


def _check_page_setup(ws, sheet_name: str) -> list[DiagnosisIssue]:
    """Check print orientation, paper size, and scaling."""
    issues: list[DiagnosisIssue] = []
    ps = ws.page_setup

    # Check if fit-to-page is configured when there's lots of data
    max_col = ws.max_column or 0
    if max_col > 10 and not ps.fitToPage:
        issues.append(DiagnosisIssue(
            type=IssueType.text_overflow,
            severity=IssueSeverity.warning,
            source=IssueSource.structural,
            description=(
                f"Sheet '{sheet_name}' has {max_col} columns but "
                f"fit-to-page is not enabled — content may overflow when printed"
            ),
            suggested_fix="set_xlsx_page_setup",
            confidence=0.7,
        ))

    return issues


def _check_margins(ws, sheet_name: str) -> list[DiagnosisIssue]:
    """Check print margins for values that are too small."""
    issues: list[DiagnosisIssue] = []
    min_margin = settings.DIAGNOSIS_MIN_MARGIN_INCHES
    pm = ws.page_margins

    sides = {
        "top": pm.top,
        "bottom": pm.bottom,
        "left": pm.left,
        "right": pm.right,
    }

    for side, value in sides.items():
        if value is not None and 0 < value < min_margin:
            issues.append(DiagnosisIssue(
                type=IssueType.margin_violation,
                severity=IssueSeverity.warning,
                source=IssueSource.structural,
                description=(
                    f"Sheet '{sheet_name}': {side} margin is {value:.2f}\" "
                    f"(minimum {min_margin}\" recommended)"
                ),
                suggested_fix="set_xlsx_margins",
                confidence=0.85,
            ))

    return issues


def _check_column_widths(ws, sheet_name: str) -> list[DiagnosisIssue]:
    """Estimate whether total column width exceeds printable area."""
    issues: list[DiagnosisIssue] = []

    # openpyxl column width is in character units (~1/7 inch per unit)
    char_width_inches = 1 / 7.0
    total_width = 0.0

    for col_letter in ws.column_dimensions:
        dim = ws.column_dimensions[col_letter]
        col_width = dim.width or 8.43  # Excel default width
        total_width += col_width * char_width_inches

    # Estimate printable width from margins
    pm = ws.page_margins
    left = pm.left or 0.75
    right = pm.right or 0.75

    # Use page setup paper size to get page width
    ps = ws.page_setup
    paper_id = ps.paperSize or 1  # default Letter
    page_w, page_h = _PAPER_SIZES.get(paper_id, (8.5, 11.0))

    # Swap for landscape
    if ps.orientation == "landscape":
        page_w, page_h = page_h, page_w

    printable_width = page_w - left - right

    if total_width > printable_width * 1.1:  # 10% tolerance
        overflow_pct = ((total_width - printable_width) / printable_width) * 100
        issues.append(DiagnosisIssue(
            type=IssueType.table_overflow,
            severity=IssueSeverity.critical,
            source=IssueSource.structural,
            description=(
                f"Sheet '{sheet_name}' content width (~{total_width:.1f}\") "
                f"exceeds printable area ({printable_width:.1f}\") by {overflow_pct:.0f}%"
            ),
            suggested_fix="set_xlsx_page_setup",
            confidence=0.75,
        ))

    return issues


def _check_fonts(ws, sheet_name: str) -> list[DiagnosisIssue]:
    """Check for uncommon fonts and small font sizes."""
    issues: list[DiagnosisIssue] = []
    min_pt = settings.DIAGNOSIS_MIN_FONT_PT
    fonts_used: set[str] = set()
    small_font_count = 0

    # Sample first 200 rows to avoid excessive processing
    for row in ws.iter_rows(max_row=200, values_only=False):
        for cell in row:
            if cell.font:
                if cell.font.name:
                    fonts_used.add(cell.font.name)
                if cell.font.size is not None and cell.font.size < min_pt:
                    if cell.value is not None:
                        small_font_count += 1

    uncommon = fonts_used - _SAFE_FONTS
    for font in uncommon:
        issues.append(DiagnosisIssue(
            type=IssueType.non_embedded_font,
            severity=IssueSeverity.warning,
            source=IssueSource.structural,
            description=(
                f"Sheet '{sheet_name}': font '{font}' may not be available on print server"
            ),
            suggested_fix="replace_font",
            confidence=0.6,
        ))

    if small_font_count:
        issues.append(DiagnosisIssue(
            type=IssueType.small_font,
            severity=IssueSeverity.warning,
            source=IssueSource.structural,
            description=(
                f"Sheet '{sheet_name}': {small_font_count} cell(s) have font size "
                f"below {min_pt}pt"
            ),
            suggested_fix="adjust_font_size",
            confidence=0.8,
        ))

    return issues


def _check_print_area(ws, sheet_name: str) -> list[DiagnosisIssue]:
    """Check if a print area is defined when the sheet has significant data."""
    issues: list[DiagnosisIssue] = []

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Only flag if the sheet has meaningful data and no print area
    if max_row > 50 and max_col > 5 and not ws.print_area:
        issues.append(DiagnosisIssue(
            type=IssueType.no_print_area,
            severity=IssueSeverity.info,
            source=IssueSource.structural,
            description=(
                f"Sheet '{sheet_name}' has {max_row} rows x {max_col} columns "
                f"but no print area defined — entire sheet will print"
            ),
            confidence=0.7,
        ))

    return issues
