"""XLSX fix tools: margins, page setup, and auto-fit."""

from __future__ import annotations

import asyncio

from app.core.log import logger
from app.schema.fix import FixResult

__all__ = (
    "adjust_xlsx_font_size",
    "auto_fit_xlsx_columns",
    "replace_xlsx_font",
    "scale_xlsx_row_heights",
    "set_xlsx_margins",
    "set_xlsx_page_setup",
    "set_xlsx_print_area",
)


# ── Font tools ───────────────────────────────────────────────────────────


async def adjust_xlsx_font_size(
    file_path: str,
    job_id: str,
    min_size_pt: float | None = None,
    max_size_pt: float | None = None,
) -> FixResult:
    """Clamp cell font sizes to a min/max range across all sheets."""
    return await asyncio.to_thread(
        _adjust_xlsx_font_size_sync,
        file_path,
        job_id,
        min_size_pt,
        max_size_pt,
    )


def _adjust_xlsx_font_size_sync(
    file_path: str,
    job_id: str,
    min_size_pt: float | None,
    max_size_pt: float | None,
) -> FixResult:
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    try:
        wb = load_workbook(file_path)
        adjusted = 0

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.font and cell.font.size is not None:
                        sz = cell.font.size
                        new_sz = sz
                        if min_size_pt is not None and sz < min_size_pt:
                            new_sz = min_size_pt
                        if max_size_pt is not None and sz > max_size_pt:
                            new_sz = max_size_pt
                        if new_sz != sz:
                            # Copy font with new size
                            f = cell.font
                            cell.font = Font(
                                name=f.name,
                                size=new_sz,
                                bold=f.bold,
                                italic=f.italic,
                                underline=f.underline,
                                strike=f.strikethrough,
                                color=f.color,
                            )
                            adjusted += 1

        wb.save(file_path)
        wb.close()

        range_desc = ""
        if min_size_pt is not None and max_size_pt is not None:
            range_desc = f"{min_size_pt}-{max_size_pt}pt"
        elif min_size_pt is not None:
            range_desc = f">={min_size_pt}pt"
        elif max_size_pt is not None:
            range_desc = f"<={max_size_pt}pt"

        return FixResult(
            tool_name="adjust_xlsx_font_size",
            job_id=job_id,
            success=True,
            description=f"Adjusted {adjusted} cell(s) to {range_desc}",
            after_value=range_desc,
        )
    except Exception as exc:
        logger.error(f"Job {job_id}: adjust_xlsx_font_size failed: {exc}")
        return FixResult(
            tool_name="adjust_xlsx_font_size",
            job_id=job_id,
            success=False,
            description=f"Failed to adjust XLSX font sizes: {exc}",
            error=str(exc),
        )


async def replace_xlsx_font(
    file_path: str,
    job_id: str,
    from_font: str,
    to_font: str,
) -> FixResult:
    """Replace all occurrences of a font with another across all sheets."""
    return await asyncio.to_thread(
        _replace_xlsx_font_sync,
        file_path,
        job_id,
        from_font,
        to_font,
    )


def _replace_xlsx_font_sync(
    file_path: str,
    job_id: str,
    from_font: str,
    to_font: str,
) -> FixResult:
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    try:
        wb = load_workbook(file_path)
        replaced = 0

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.font and cell.font.name == from_font:
                        f = cell.font
                        cell.font = Font(
                            name=to_font,
                            size=f.size,
                            bold=f.bold,
                            italic=f.italic,
                            underline=f.underline,
                            strike=f.strikethrough,
                            color=f.color,
                        )
                        replaced += 1

        wb.save(file_path)
        wb.close()

        return FixResult(
            tool_name="replace_xlsx_font",
            job_id=job_id,
            success=True,
            description=f"Replaced '{from_font}' with '{to_font}' in {replaced} cell(s)",
            before_value=from_font,
            after_value=to_font,
        )
    except Exception as exc:
        logger.error(f"Job {job_id}: replace_xlsx_font failed: {exc}")
        return FixResult(
            tool_name="replace_xlsx_font",
            job_id=job_id,
            success=False,
            description=f"Failed to replace XLSX font: {exc}",
            error=str(exc),
        )


# ── Print area ───────────────────────────────────────────────────────────


async def set_xlsx_print_area(
    file_path: str,
    job_id: str,
    area: str | None = None,
) -> FixResult:
    """Set print area on sheets that lack one.

    If area is None, auto-detects the used range per sheet.
    area format example: 'A1:M50' (applied to all sheets without a print area).
    """
    return await asyncio.to_thread(
        _set_xlsx_print_area_sync,
        file_path,
        job_id,
        area,
    )


def _set_xlsx_print_area_sync(
    file_path: str,
    job_id: str,
    area: str | None,
) -> FixResult:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    try:
        wb = load_workbook(file_path)
        sheets_set = 0
        details: list[str] = []

        for ws in wb.worksheets:
            # Skip sheets that already have a print area
            if ws.print_area:
                continue

            if area:
                pa = area
            else:
                # Auto-detect used range
                max_row = ws.max_row or 1
                max_col = ws.max_column or 1
                if max_row <= 1 and max_col <= 1:
                    continue  # Empty sheet
                end_col = get_column_letter(max_col)
                pa = f"A1:{end_col}{max_row}"

            ws.print_area = pa
            sheets_set += 1
            details.append(f"'{ws.title}': {pa}")

        wb.save(file_path)
        wb.close()

        return FixResult(
            tool_name="set_xlsx_print_area",
            job_id=job_id,
            success=True,
            description=(
                f"Set print area on {sheets_set} sheet(s): " + "; ".join(details)
                if details
                else "All sheets already have print areas"
            ),
            after_value="; ".join(details) if details else "no change",
        )
    except Exception as exc:
        logger.error(f"Job {job_id}: set_xlsx_print_area failed: {exc}")
        return FixResult(
            tool_name="set_xlsx_print_area",
            job_id=job_id,
            success=False,
            description=f"Failed to set XLSX print area: {exc}",
            error=str(exc),
        )


# ── Row heights ──────────────────────────────────────────────────────────


async def scale_xlsx_row_heights(
    file_path: str,
    job_id: str,
    auto_fit: bool = True,
) -> FixResult:
    """Auto-fit row heights to cell content across all sheets.

    Estimates optimal row height based on font size and text wrapping.
    """
    return await asyncio.to_thread(
        _scale_xlsx_row_heights_sync,
        file_path,
        job_id,
        auto_fit,
    )


def _scale_xlsx_row_heights_sync(
    file_path: str,
    job_id: str,
    auto_fit: bool,
) -> FixResult:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_path)
        rows_adjusted = 0

        for ws in wb.worksheets:
            for row_idx in range(1, (ws.max_row or 0) + 1):
                rd = ws.row_dimensions[row_idx]
                if not auto_fit:
                    # Just clear custom heights to let Excel auto-fit
                    if rd.customHeight:
                        rd.customHeight = False
                        rd.height = None
                        rows_adjusted += 1
                    continue

                # Estimate optimal height from cell content
                max_height = 15.0  # default row height in points
                for cell in ws[row_idx]:
                    if cell.value is None:
                        continue
                    font_size = cell.font.size if cell.font and cell.font.size else 11
                    text = str(cell.value)
                    lines = text.count("\n") + 1

                    # If text wrapping is enabled, estimate wrapped lines
                    if cell.alignment and cell.alignment.wrap_text:
                        col_width = ws.column_dimensions[cell.column_letter].width or 8.43
                        chars_per_line = max(1, int(col_width * 1.2))
                        lines = max(lines, len(text) // chars_per_line + 1)

                    estimated_height = lines * (font_size * 1.4 + 2)
                    max_height = max(max_height, estimated_height)

                if max_height > (rd.height or 15.0):
                    rd.height = max_height
                    rows_adjusted += 1

        wb.save(file_path)
        wb.close()

        return FixResult(
            tool_name="scale_xlsx_row_heights",
            job_id=job_id,
            success=True,
            description=f"Adjusted {rows_adjusted} row(s) across all sheets",
            after_value=f"{rows_adjusted} rows adjusted",
        )
    except Exception as exc:
        logger.error(f"Job {job_id}: scale_xlsx_row_heights failed: {exc}")
        return FixResult(
            tool_name="scale_xlsx_row_heights",
            job_id=job_id,
            success=False,
            description=f"Failed to scale XLSX row heights: {exc}",
            error=str(exc),
        )


# ── Margin tools ─────────────────────────────────────────────────────────


async def set_xlsx_margins(
    file_path: str,
    job_id: str,
    top: float = 0.75,
    bottom: float = 0.75,
    left: float = 0.75,
    right: float = 0.75,
) -> FixResult:
    """Set print margins on all sheets. Values in inches."""
    return await asyncio.to_thread(
        _set_xlsx_margins_sync,
        file_path,
        job_id,
        top,
        bottom,
        left,
        right,
    )


def _set_xlsx_margins_sync(
    file_path: str,
    job_id: str,
    top: float,
    bottom: float,
    left: float,
    right: float,
) -> FixResult:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_path)
        sheets_changed = 0

        for ws in wb.worksheets:
            pm = ws.page_margins
            changed = False
            for attr, val in [("top", top), ("bottom", bottom), ("left", left), ("right", right)]:
                if getattr(pm, attr) != val:
                    setattr(pm, attr, val)
                    changed = True
            if changed:
                sheets_changed += 1

        wb.save(file_path)
        wb.close()

        return FixResult(
            tool_name="set_xlsx_margins",
            job_id=job_id,
            success=True,
            description=f"Set margins on {sheets_changed} sheet(s) to T={top} B={bottom} L={left} R={right} inches",
            after_value=f"T={top} B={bottom} L={left} R={right}",
        )
    except Exception as exc:
        logger.error(f"Job {job_id}: set_xlsx_margins failed: {exc}")
        return FixResult(
            tool_name="set_xlsx_margins",
            job_id=job_id,
            success=False,
            description=f"Failed to set XLSX margins: {exc}",
            error=str(exc),
        )


async def set_xlsx_page_setup(
    file_path: str,
    job_id: str,
    orientation: str = "portrait",
    paper_size: int = 1,
    fit_to_page: bool = True,
) -> FixResult:
    """Configure print page setup on all sheets.

    Args:
        orientation: 'portrait' or 'landscape'
        paper_size: 1=Letter, 9=A4
        fit_to_page: if True, fit all columns to one page width
    """
    return await asyncio.to_thread(
        _set_xlsx_page_setup_sync,
        file_path,
        job_id,
        orientation,
        paper_size,
        fit_to_page,
    )


def _set_xlsx_page_setup_sync(
    file_path: str,
    job_id: str,
    orientation: str,
    paper_size: int,
    fit_to_page: bool,
) -> FixResult:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_path)
        sheets_changed = 0

        for ws in wb.worksheets:
            ps = ws.page_setup
            ps.orientation = orientation
            ps.paperSize = paper_size

            if fit_to_page:
                ps.fitToPage = True
                ps.fitToWidth = 1
                ps.fitToHeight = 0  # 0 = auto (fit width, allow multiple pages tall)

            sheets_changed += 1

        wb.save(file_path)
        wb.close()

        paper_name = {1: "Letter", 9: "A4"}.get(paper_size, str(paper_size))
        return FixResult(
            tool_name="set_xlsx_page_setup",
            job_id=job_id,
            success=True,
            description=(
                f"Set page setup on {sheets_changed} sheet(s): "
                f"{orientation}, {paper_name}" + (", fit-to-page enabled" if fit_to_page else "")
            ),
            after_value=f"{orientation}, {paper_name}, fit={fit_to_page}",
        )
    except Exception as exc:
        logger.error(f"Job {job_id}: set_xlsx_page_setup failed: {exc}")
        return FixResult(
            tool_name="set_xlsx_page_setup",
            job_id=job_id,
            success=False,
            description=f"Failed to set XLSX page setup: {exc}",
            error=str(exc),
        )


async def auto_fit_xlsx_columns(
    file_path: str,
    job_id: str,
    max_col_width: float = 30.0,
    min_col_width: float = 5.0,
    shrink_margins: bool = True,
) -> FixResult:
    """Analyse content widths across all sheets and auto-size columns so the
    table fits within a single page width.  When *shrink_margins* is True the
    page margins are also tightened to 0.4" to maximise printable area.

    The function automatically picks portrait or landscape orientation
    depending on which one allows the columns to print without scaling, and
    enables fit-to-page as a safety net.
    """
    return await asyncio.to_thread(
        _auto_fit_xlsx_columns_sync,
        file_path,
        job_id,
        max_col_width,
        min_col_width,
        shrink_margins,
    )


# Standard paper sizes in inches (width, height) portrait
_PAPER_SIZES: dict[int, tuple[float, float]] = {
    1: (8.5, 11.0),  # Letter
    9: (8.27, 11.69),  # A4
}


def _auto_fit_xlsx_columns_sync(
    file_path: str,
    job_id: str,
    max_col_width: float,
    min_col_width: float,
    shrink_margins: bool,
) -> FixResult:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    try:
        wb = load_workbook(file_path)
        sheets_changed = 0
        details: list[str] = []

        for ws in wb.worksheets:
            max_col = ws.max_column or 0
            if max_col == 0:
                continue

            # ── 1. Tighten margins if requested ──────────────────────
            margin = 0.4  # inches
            if shrink_margins:
                pm = ws.page_margins
                pm.top = margin
                pm.bottom = margin
                pm.left = margin
                pm.right = margin

            # ── 2. Measure optimal column widths from cell content ───
            optimal_widths: dict[int, float] = {}
            for col_idx in range(1, max_col + 1):
                best = min_col_width
                for row in ws.iter_rows(
                    min_col=col_idx,
                    max_col=col_idx,
                    max_row=min(ws.max_row or 0, 500),
                    values_only=False,
                ):
                    cell = row[0]
                    if cell.value is not None:
                        text = str(cell.value)
                        # Approximate character width (openpyxl units ≈ chars)
                        cell_width = len(text) + 2  # small padding
                        if cell.font and cell.font.bold:
                            cell_width *= 1.1
                        best = max(best, min(cell_width, max_col_width))
                optimal_widths[col_idx] = best

            total_content_chars = sum(optimal_widths.values())

            # ── 3. Choose orientation ────────────────────────────────
            ps = ws.page_setup
            paper_id = ps.paperSize or 1
            page_w_p, page_h_p = _PAPER_SIZES.get(paper_id, (8.5, 11.0))

            left_m = margin if shrink_margins else (ws.page_margins.left or 0.75)
            right_m = margin if shrink_margins else (ws.page_margins.right or 0.75)

            printable_portrait = page_w_p - left_m - right_m
            printable_landscape = page_h_p - left_m - right_m

            # openpyxl column width 1 unit ≈ 1/7 inch
            char_to_inch = 1 / 7.0
            total_content_inches = total_content_chars * char_to_inch

            if total_content_inches <= printable_portrait:
                chosen_orientation = "portrait"
                printable = printable_portrait
            else:
                chosen_orientation = "landscape"
                printable = printable_landscape

            ps.orientation = chosen_orientation
            if chosen_orientation == "landscape":
                # openpyxl doesn't swap for us, so just set orientation
                pass

            # ── 4. Scale columns to fit printable width ──────────────
            printable_chars = printable / char_to_inch
            if total_content_chars > printable_chars:
                scale = printable_chars / total_content_chars
                for col_idx in optimal_widths:
                    optimal_widths[col_idx] = max(
                        min_col_width,
                        optimal_widths[col_idx] * scale,
                    )

            # Apply widths
            for col_idx, w in optimal_widths.items():
                letter = get_column_letter(col_idx)
                ws.column_dimensions[letter].width = w

            # ── 5. Enable fit-to-page as safety net ──────────────────
            ps.fitToPage = True
            ps.fitToWidth = 1
            ps.fitToHeight = 0  # auto height

            sheets_changed += 1
            details.append(f"'{ws.title}': {max_col} cols, {chosen_orientation}, margins {margin}\"")

        wb.save(file_path)
        wb.close()

        return FixResult(
            tool_name="auto_fit_xlsx_columns",
            job_id=job_id,
            success=True,
            description=(f"Auto-fitted columns on {sheets_changed} sheet(s): " + "; ".join(details)),
            after_value="; ".join(details),
        )
    except Exception as exc:
        logger.error(f"Job {job_id}: auto_fit_xlsx_columns failed: {exc}")
        return FixResult(
            tool_name="auto_fit_xlsx_columns",
            job_id=job_id,
            success=False,
            description=f"Failed to auto-fit XLSX columns: {exc}",
            error=str(exc),
        )
